import webview
import os
import sys
import json
import uuid
import hashlib
import platform
import datetime
import base64
import socket
import struct
import time
import sqlite3

from cryptography.hazmat.primitives.asymmetric import padding
from cryptography.hazmat.primitives import hashes, serialization
from cryptography.hazmat.backends import default_backend

try:
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    _OPENPYXL_OK = True
except ImportError:
    _OPENPYXL_OK = False

# ─── paths ────────────────────────────────────────────────────────────────────
def resource_path(filename):
    if hasattr(sys, '_MEIPASS'):
        return os.path.join(sys._MEIPASS, filename)
    return os.path.join(os.path.dirname(os.path.abspath(__file__)), filename)

# ── Hidden data folder (sits next to the app, invisible in file explorer) ──────
_APP_DIR     = os.path.dirname(os.path.abspath(__file__))
_DATA_DIR    = os.path.join(_APP_DIR, '.hbillsoft')
os.makedirs(_DATA_DIR, exist_ok=True)

# On Windows, mark the folder as hidden via attrib
if platform.system() == 'Windows':
    try:
        import subprocess
        subprocess.call(['attrib', '+H', _DATA_DIR], shell=False)
    except Exception:
        pass

def data_path(filename):
    """Resolve a data file into the hidden .hbillsoft folder."""
    return os.path.join(_DATA_DIR, filename)

CONFIG_FILE    = data_path('config.json')
LICENSE_FILE   = data_path('license.dat')
LASTRUN_FILE   = data_path('lastrun.dat')   # stores last-seen date (hidden from user)
EXCEL_DB_FILE  = data_path('sales_data.xlsx')  # auto-saved sales database
DB_FILE        = data_path('sales_data.db')    # SQLite database

# ─── RSA public key (vendor embeds this; private key stays with vendor) ───────
# NOTE: When you generate a new key pair in keygen_gui.py (tkinter),
#       replace the block below with the new public key shown in that dialog.
PUBLIC_KEY_PEM = b"""-----BEGIN PUBLIC KEY-----
MIIBIjANBgkqhkiG9w0BAQEFAAOCAQ8AMIIBCgKCAQEA8N47o0fqzoqxp/bYRokm
sNyT15U7zOc5v/dK7d18ebiblOpTnEpAPEqwg4id/DeAEJyE+oizP8VO2hE21GOo
qTayv6/K9nWOsz17fgzEIpqAAwN8tmeNFrnfOb6UDJ+s4UZPuE3Y8E8hxgTIZWfZ
FA97+tw55p05FwXwxunnQXhGNMPFJppz4iDk9hnZSzIPagj53PMjg/VFcH21I2Yz
8Ey5UWU/P25KnknLWNhwZ3xHrIwJcf29YFxeIE5sEfK8F+4SSx3Te7cI/UwZNWqS
xcMjdssJaEu69EqNbQbYti3SmSAfyNQh1P6VHKVI3jGhLRdvEY7Ml7JWWlfHvD5o
swIDAQAB
-----END PUBLIC KEY-----"""

# ─── Clock tamper detection ───────────────────────────────────────────────────

# Max allowed drift between system clock and NTP (seconds)
NTP_DRIFT_TOLERANCE = 86400          # 1 day — generous for offline use
# NTP servers to try (in order)
NTP_SERVERS = [
    'time.cloudflare.com',
    'time.google.com',
    'pool.ntp.org',
    'time.windows.com',
]

def _get_ntp_time() -> datetime.date | None:
    """
    Query NTP servers for the real current time.
    Returns a date if reachable, None if all servers fail (offline).
    Uses raw UDP — no external libraries needed.
    """
    NTP_DELTA = 2208988800  # seconds between 1900-01-01 and 1970-01-01
    for server in NTP_SERVERS:
        try:
            client = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
            client.settimeout(3)
            data = b'\x1b' + 47 * b'\0'
            client.sendto(data, (server, 123))
            response, _ = client.recvfrom(1024)
            client.close()
            # Transmit timestamp is at bytes 40-47
            tx_timestamp = struct.unpack('!I', response[40:44])[0]
            ntp_epoch    = tx_timestamp - NTP_DELTA
            return datetime.date.fromtimestamp(ntp_epoch)
        except Exception:
            continue
    return None  # all servers unreachable (offline)


def _load_last_run_date() -> datetime.date | None:
    """Load the last-seen date stored on disk (obfuscated, not plain text)."""
    try:
        with open(LASTRUN_FILE, 'r') as f:
            raw = f.read().strip()
        # Decode: base64 → reverse → date string
        decoded = base64.b64decode(raw.encode()).decode()[::-1]
        return datetime.date.fromisoformat(decoded)
    except Exception:
        return None


def _save_last_run_date(date: datetime.date):
    """Persist today's date to disk (lightly obfuscated)."""
    try:
        # Obfuscate: reverse string → base64 (not encryption, just not plain-text)
        raw     = date.isoformat()[::-1]
        encoded = base64.b64encode(raw.encode()).decode()
        with open(LASTRUN_FILE, 'w') as f:
            f.write(encoded)
    except Exception:
        pass  # read-only filesystem edge case — don't crash


def detect_clock_tampering(issued_date_str: str) -> dict:
    """
    Run all three clock-tamper checks.
    Returns:
      tampered (bool)  — True if tampering detected
      reason   (str)   — human-readable reason if tampered
      source   (str)   — 'ntp' | 'lastrun' | 'issued' | 'ok'
    """
    today = datetime.date.today()

    # ── Layer 1: Issued-date check (always offline) ───────────────────────────
    # If today is BEFORE the issue date, the clock was clearly rolled back.
    try:
        issued = datetime.date.fromisoformat(issued_date_str)
        if today < issued:
            return {
                'tampered': True,
                'reason':   'System date is before the license issue date. Please correct your system clock.',
                'source':   'issued'
            }
    except Exception:
        pass

    # ── Layer 2: Last-run date check (always offline) ─────────────────────────
    # If today is BEFORE the last time the app ran, the clock was rolled back.
    last_run = _load_last_run_date()
    if last_run and today < last_run:
        return {
            'tampered': True,
            'reason':   'System date appears to have been set back. Please correct your system clock.',
            'source':   'lastrun'
        }

    # ── Layer 3: NTP check (online, best-effort) ──────────────────────────────
    # If we can reach a time server, compare against system clock.
    ntp_date = _get_ntp_time()
    if ntp_date is not None:
        drift_days = abs((today - ntp_date).days)
        if drift_days > 1:   # more than 1 day off
            return {
                'tampered': True,
                'reason':   f'System clock is {drift_days} day(s) off from internet time. Please correct your system clock.',
                'source':   'ntp'
            }

    # All checks passed — update last-run date
    _save_last_run_date(today)
    return {'tampered': False, 'reason': '', 'source': 'ok'}


# ─── System Code (machine fingerprint) ────────────────────────────────────────
def get_system_code() -> str:
    """Build a stable machine fingerprint from hardware identifiers."""
    parts = []

    try:
        parts.append(str(uuid.getnode()))
    except Exception:
        pass
    try:
        parts.append(platform.node())
    except Exception:
        pass
    try:
        parts.append(platform.machine())
    except Exception:
        pass
    try:
        parts.append(platform.processor())
    except Exception:
        pass

    for path in ['/etc/machine-id', '/var/lib/dbus/machine-id']:
        try:
            with open(path) as f:
                parts.append(f.read().strip())
            break
        except Exception:
            pass
    if platform.system() == 'Windows':
        try:
            import winreg
            key = winreg.OpenKey(winreg.HKEY_LOCAL_MACHINE,
                                 r'SOFTWARE\Microsoft\Cryptography')
            val, _ = winreg.QueryValueEx(key, 'MachineGuid')
            parts.append(val)
        except Exception:
            pass

    raw    = '|'.join(parts)
    digest = hashlib.sha256(raw.encode()).hexdigest()
    d      = digest[:16].upper()
    return '-'.join([d[i:i+4] for i in range(0, 16, 4)])


# ─── License validation ───────────────────────────────────────────────────────
def _load_public_key():
    return serialization.load_pem_public_key(PUBLIC_KEY_PEM, backend=default_backend())

def verify_license_key(license_key: str, system_code: str) -> dict:
    """
    Returns dict with keys:
      valid (bool), expired (bool), expiry (str), customer (str),
      issued (str), days_left (int), error (str), clock_tampered (bool)
    """
    try:
        # Sanitize
        clean_key = license_key.strip()
        clean_key = clean_key.replace('\r', '').replace('\n', '').replace(' ', '')
        clean_key = clean_key.lstrip('\ufeff')

        raw         = base64.b64decode(clean_key.encode())
        bundle      = json.loads(raw)
        payload_b64 = bundle['payload']
        sig_b64     = bundle['signature']

        payload_bytes = base64.b64decode(payload_b64)
        sig_bytes     = base64.b64decode(sig_b64)

        # 1. Verify RSA signature
        pub_key = _load_public_key()
        pub_key.verify(sig_bytes, payload_bytes, padding.PKCS1v15(), hashes.SHA256())

        # 2. Decode payload
        data = json.loads(payload_bytes)

        # 3. Check system code binding
        if data.get('system_code', '').upper() != system_code.upper():
            return {'valid': False, 'error': 'License is not valid for this machine.', 'clock_tampered': False}

        # 4. Check product
        if data.get('product', '') != 'HBILLSOFT':
            return {'valid': False, 'error': 'Invalid product in license.', 'clock_tampered': False}

        # 5. Clock tamper detection (all three layers)
        issued_str  = data.get('issued', datetime.date.today().isoformat())
        tamper_info = detect_clock_tampering(issued_str)
        if tamper_info['tampered']:
            return {
                'valid':          False,
                'expired':        False,
                'clock_tampered': True,
                'error':          tamper_info['reason'],
                'expiry':         data.get('expiry', ''),
                'customer':       data.get('customer', ''),
                'issued':         issued_str,
                'days_left':      0,
            }

        # 6. Check expiry
        expiry_date = datetime.date.fromisoformat(data['expiry'])
        today       = datetime.date.today()
        expired     = today > expiry_date
        days_left   = (expiry_date - today).days if not expired else 0

        return {
            'valid':          True,
            'expired':        expired,
            'expiry':         data['expiry'],
            'customer':       data.get('customer', ''),
            'issued':         issued_str,
            'days_left':      days_left,
            'clock_tampered': False,
            'error':          ''
        }
    except Exception as e:
        return {'valid': False, 'error': f'License verification failed: {e}', 'clock_tampered': False}


def load_saved_license() -> str:
    try:
        with open(LICENSE_FILE, 'r') as f:
            return f.read().strip()
    except Exception:
        return ''

def save_license(key: str):
    with open(LICENSE_FILE, 'w') as f:
        f.write(key.strip())

# ─── Config helpers ───────────────────────────────────────────────────────────
def load_config_from_file():
    try:
        with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return {}

def save_config_to_file(data):
    try:
        with open(CONFIG_FILE, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=4, ensure_ascii=False)
        return True
    except Exception:
        return False

# ─── SQLite Database Functions ────────────────────────────────────────────────
def init_database():
    """Initialize SQLite database with orders, items, menu and categories tables."""
    try:
        conn = sqlite3.connect(DB_FILE)
        cursor = conn.cursor()
        
        # Create orders table
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS orders (
                id TEXT PRIMARY KEY,
                date TEXT NOT NULL,
                subtotal REAL NOT NULL,
                discount REAL NOT NULL,
                sgst REAL NOT NULL,
                cgst REAL NOT NULL,
                total REAL NOT NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        
        # Create order_items table (links items to orders)
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS order_items (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                order_id TEXT NOT NULL,
                item_name TEXT NOT NULL,
                item_id TEXT,
                quantity INTEGER NOT NULL,
                price REAL NOT NULL,
                item_total REAL NOT NULL,
                FOREIGN KEY(order_id) REFERENCES orders(id)
            )
        ''')

        # Create categories table
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS categories (
                id TEXT PRIMARY KEY,
                name TEXT NOT NULL,
                icon TEXT NOT NULL DEFAULT '🍽️',
                sort_order INTEGER DEFAULT 0
            )
        ''')

        # Create menu_items table
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS menu_items (
                id INTEGER PRIMARY KEY,
                name TEXT NOT NULL,
                price REAL NOT NULL,
                category TEXT NOT NULL,
                image TEXT DEFAULT '🍽️',
                image_data TEXT,
                sort_order INTEGER DEFAULT 0
            )
        ''')

        # Create settings table (single JSON row)
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS app_settings (
                id   INTEGER PRIMARY KEY CHECK (id = 1),
                data TEXT NOT NULL
            )
        ''')

        # Create cart table (persists current in-progress cart)
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS cart (
                id   INTEGER PRIMARY KEY CHECK (id = 1),
                data TEXT NOT NULL
            )
        ''')

        conn.commit()
        conn.close()
        return True
    except Exception as e:
        print(f"Database init error: {e}")
        return False

# ─── Menu & Categories DB helpers ─────────────────────────────────────────────

def save_categories_to_db(categories: list) -> dict:
    """Replace all categories in the DB."""
    try:
        conn = sqlite3.connect(DB_FILE)
        cursor = conn.cursor()
        cursor.execute('DELETE FROM categories')
        for i, cat in enumerate(categories):
            cursor.execute('''
                INSERT OR REPLACE INTO categories (id, name, icon, sort_order)
                VALUES (?, ?, ?, ?)
            ''', (cat.get('id', ''), cat.get('name', ''), cat.get('icon', '🍽️'), i))
        conn.commit()
        conn.close()
        return {'ok': True}
    except Exception as e:
        return {'ok': False, 'error': str(e)}

def load_categories_from_db() -> dict:
    """Load all categories from the DB."""
    try:
        conn = sqlite3.connect(DB_FILE)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        cursor.execute('SELECT id, name, icon FROM categories ORDER BY sort_order, rowid')
        rows = cursor.fetchall()
        conn.close()
        categories = [{'id': r['id'], 'name': r['name'], 'icon': r['icon']} for r in rows]
        return {'ok': True, 'categories': categories}
    except Exception as e:
        return {'ok': False, 'error': str(e), 'categories': []}

def save_menu_to_db(menu: list) -> dict:
    """Replace all menu items in the DB."""
    try:
        conn = sqlite3.connect(DB_FILE)
        cursor = conn.cursor()
        cursor.execute('DELETE FROM menu_items')
        for i, item in enumerate(menu):
            cursor.execute('''
                INSERT OR REPLACE INTO menu_items (id, name, price, category, image, image_data, sort_order)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            ''', (
                item.get('id'),
                item.get('name', ''),
                float(item.get('price', 0)),
                item.get('category', ''),
                item.get('image', '🍽️'),
                item.get('imageData', None),
                i
            ))
        conn.commit()
        conn.close()
        return {'ok': True}
    except Exception as e:
        return {'ok': False, 'error': str(e)}

def load_menu_from_db() -> dict:
    """Load all menu items from the DB."""
    try:
        conn = sqlite3.connect(DB_FILE)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        cursor.execute('SELECT id, name, price, category, image, image_data FROM menu_items ORDER BY sort_order, id')
        rows = cursor.fetchall()
        conn.close()
        menu = []
        for r in rows:
            item = {
                'id': r['id'],
                'name': r['name'],
                'price': r['price'],
                'category': r['category'],
                'image': r['image']
            }
            if r['image_data']:
                item['imageData'] = r['image_data']
            menu.append(item)
        return {'ok': True, 'menu': menu}
    except Exception as e:
        return {'ok': False, 'error': str(e), 'menu': []}

def save_settings_to_db(settings: dict) -> dict:
    """Persist app settings as a single JSON blob."""
    try:
        conn = sqlite3.connect(DB_FILE)
        cursor = conn.cursor()
        # Strip internal-only keys before saving
        clean = {k: v for k, v in settings.items() if k != 'sessionCount'}
        cursor.execute(
            'INSERT OR REPLACE INTO app_settings (id, data) VALUES (1, ?)',
            (json.dumps(clean, ensure_ascii=False),)
        )
        conn.commit()
        conn.close()
        return {'ok': True}
    except Exception as e:
        return {'ok': False, 'error': str(e)}

def load_settings_from_db() -> dict:
    """Load app settings from SQLite. Returns {} if not yet saved."""
    try:
        conn = sqlite3.connect(DB_FILE)
        cursor = conn.cursor()
        cursor.execute('SELECT data FROM app_settings WHERE id = 1')
        row = cursor.fetchone()
        conn.close()
        if row:
            return {'ok': True, 'settings': json.loads(row[0])}
        return {'ok': True, 'settings': {}}
    except Exception as e:
        return {'ok': False, 'error': str(e), 'settings': {}}

def save_cart_to_db(cart: list) -> dict:
    """Persist the current cart as a single JSON blob."""
    try:
        conn = sqlite3.connect(DB_FILE)
        cursor = conn.cursor()
        cursor.execute(
            'INSERT OR REPLACE INTO cart (id, data) VALUES (1, ?)',
            (json.dumps(cart, ensure_ascii=False),)
        )
        conn.commit()
        conn.close()
        return {'ok': True}
    except Exception as e:
        return {'ok': False, 'error': str(e)}

def load_cart_from_db() -> dict:
    """Load the persisted cart from SQLite. Returns [] if empty."""
    try:
        conn = sqlite3.connect(DB_FILE)
        cursor = conn.cursor()
        cursor.execute('SELECT data FROM cart WHERE id = 1')
        row = cursor.fetchone()
        conn.close()
        if row:
            return {'ok': True, 'cart': json.loads(row[0])}
        return {'ok': True, 'cart': []}
    except Exception as e:
        return {'ok': False, 'error': str(e), 'cart': []}

def save_order_to_db(order: dict) -> dict:
    """Save an order and its items to SQLite database."""
    try:
        init_database()
        conn = sqlite3.connect(DB_FILE)
        cursor = conn.cursor()
        
        order_id = order.get('id', str(uuid.uuid4()))
        date = order.get('date', datetime.datetime.now().isoformat())
        subtotal = float(order.get('subtotal', 0))
        discount = float(order.get('discount', 0))
        sgst = float(order.get('sgst', 0))
        cgst = float(order.get('cgst', 0))
        total = float(order.get('total', 0))
        
        # Save order
        cursor.execute('''
            INSERT OR REPLACE INTO orders 
            (id, date, subtotal, discount, sgst, cgst, total)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        ''', (order_id, date, subtotal, discount, sgst, cgst, total))
        
        # Save order items
        for item in order.get('items', []):
            cursor.execute('''
                INSERT INTO order_items 
                (order_id, item_name, item_id, quantity, price, item_total)
                VALUES (?, ?, ?, ?, ?, ?)
            ''', (
                order_id,
                item.get('name', ''),
                item.get('id', ''),
                item.get('qty', 1),
                item.get('price', 0),
                float(item.get('qty', 1)) * float(item.get('price', 0))
            ))
        
        conn.commit()
        conn.close()
        return {'ok': True, 'order_id': order_id}
    except Exception as e:
        print(f"Error saving order to DB: {e}")
        return {'ok': False, 'error': str(e)}

def load_orders_from_db() -> dict:
    """Load all orders from SQLite database."""
    try:
        init_database()
        if not os.path.exists(DB_FILE):
            return {'ok': True, 'orders': []}
        
        conn = sqlite3.connect(DB_FILE)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        
        # Get all orders
        cursor.execute('SELECT * FROM orders ORDER BY date DESC')
        orders_rows = cursor.fetchall()
        
        orders = []
        for order_row in orders_rows:
            order_id = order_row['id']
            
            # Get items for this order
            cursor.execute('''
                SELECT item_name, item_id, quantity, price, item_total 
                FROM order_items 
                WHERE order_id = ?
            ''', (order_id,))
            items_rows = cursor.fetchall()
            
            items = [
                {
                    'name': item['item_name'],
                    'id': item['item_id'],
                    'qty': item['quantity'],
                    'price': item['price'],
                    'total': item['item_total']
                }
                for item in items_rows
            ]
            
            order = {
                'id': order_row['id'],
                'date': order_row['date'],
                'items': items,
                'subtotal': order_row['subtotal'],
                'discount': order_row['discount'],
                'sgst': order_row['sgst'],
                'cgst': order_row['cgst'],
                'gst': order_row['sgst'],  # For compatibility
                'total': order_row['total']
            }
            orders.append(order)
        
        conn.close()
        return {'ok': True, 'orders': orders}
    except Exception as e:
        print(f"Error loading orders from DB: {e}")
        return {'ok': False, 'error': str(e), 'orders': []}

def get_sales_summary(from_date=None, to_date=None) -> dict:
    """
    Get sales summary with items grouped by name, quantities, and date-based filtering.
    Returns items with total quantity sold and total revenue.
    """
    try:
        init_database()
        if not os.path.exists(DB_FILE):
            return {'ok': True, 'summary': [], 'total_revenue': 0}
        
        conn = sqlite3.connect(DB_FILE)
        cursor = conn.cursor()
        
        # Build query with optional date filtering
        query = '''
            SELECT 
                oi.item_name,
                SUM(oi.quantity) as total_qty,
                AVG(oi.price) as avg_price,
                SUM(oi.item_total) as total_revenue,
                COUNT(DISTINCT oi.order_id) as order_count
            FROM order_items oi
            JOIN orders o ON oi.order_id = o.id
        '''
        
        params = []
        if from_date or to_date:
            query += ' WHERE'
            if from_date:
                query += ' o.date >= ?'
                params.append(from_date)
            if from_date and to_date:
                query += ' AND'
            if to_date:
                query += ' o.date <= ?'
                params.append(to_date)
        
        query += ' GROUP BY oi.item_name ORDER BY total_revenue DESC'
        
        cursor.execute(query, params)
        rows = cursor.fetchall()
        
        summary = []
        total_revenue = 0
        for row in rows:
            item_summary = {
                'name': row[0],
                'quantity': int(row[1]),
                'avg_price': round(row[2], 2),
                'revenue': round(row[3], 2),
                'orders': int(row[4])
            }
            summary.append(item_summary)
            total_revenue += row[3]
        # Also compute total distinct orders and tax totals in the filtered range
        orders_count = None
        total_sgst = 0.0
        total_cgst = 0.0
        grand_total_including_tax = 0.0
        try:
            # Count distinct orders
            count_query = 'SELECT COUNT(DISTINCT o.id) FROM orders o'
            sum_query = 'SELECT SUM(o.sgst), SUM(o.cgst), SUM(o.total) FROM orders o'
            if from_date or to_date:
                where_parts = []
                params_count = []
                if from_date:
                    where_parts.append('o.date >= ?')
                    params_count.append(from_date)
                if to_date:
                    where_parts.append('o.date <= ?')
                    params_count.append(to_date)
                where_clause = ' WHERE ' + ' AND '.join(where_parts)
                cursor.execute(count_query + where_clause, params_count)
                orders_count = cursor.fetchone()[0]
                cursor.execute(sum_query + where_clause, params_count)
            else:
                cursor.execute(count_query)
                orders_count = cursor.fetchone()[0]
                cursor.execute(sum_query)

            sums = cursor.fetchone()
            if sums:
                total_sgst = float(sums[0] or 0)
                total_cgst = float(sums[1] or 0)
                grand_total_including_tax = float(sums[2] or 0)
        except Exception:
            orders_count = None

        conn.close()
        return {
            'ok': True,
            'summary': summary,
            'total_revenue': round(total_revenue, 2),            # item totals (pre-tax)
            'orders_count': orders_count,
            'total_sgst': round(total_sgst, 2),
            'total_cgst': round(total_cgst, 2),
            'total_gst': round(total_sgst + total_cgst, 2),
            'grand_total': round(grand_total_including_tax, 2)   # including tax
        }
    except Exception as e:
        print(f"Error getting sales summary: {e}")
        return {'ok': False, 'error': str(e), 'summary': [], 'total_revenue': 0}

# ─── Excel report helpers ────────────────────────────────────────────────────

def _get_daily_breakdown(from_date=None, to_date=None) -> list:
    """
    Return a list of day dicts ordered by date:
      [{ 'date': datetime.date, 'items': [{name, category, quantity, price, amount}],
         'subtotal': float, 'sgst': float, 'cgst': float, 'day_total': float }, ...]
    Items within a day are aggregated by (item_name, price) so each unique
    name+rate combination gets its own numbered row, matching the receipt layout.
    """
    try:
        init_database()
        if not os.path.exists(DB_FILE):
            return []
        conn = sqlite3.connect(DB_FILE)
        cur  = conn.cursor()

        # Build WHERE clause
        params = []
        where  = ''
        if from_date or to_date:
            parts = []
            if from_date:
                parts.append('o.date >= ?'); params.append(from_date)
            if to_date:
                parts.append('o.date <= ?'); params.append(to_date)
            where = 'WHERE ' + ' AND '.join(parts)

        # Per-day tax & total from orders table
        cur.execute(f'''
            SELECT DATE(o.date) as day,
                   SUM(o.sgst)  as sgst,
                   SUM(o.cgst)  as cgst,
                   SUM(o.total) as day_total
            FROM orders o {where}
            GROUP BY day ORDER BY day
        ''', params)
        day_rows = cur.fetchall()

        # Per-day item detail (join menu_items for category)
        cur.execute(f'''
            SELECT DATE(o.date)    as day,
                   oi.item_name,
                   COALESCE(m.category, '') as category,
                   SUM(oi.quantity)          as qty,
                   oi.price,
                   SUM(oi.item_total)        as amount
            FROM order_items oi
            JOIN orders o ON oi.order_id = o.id
            LEFT JOIN menu_items m ON oi.item_name = m.name
            {where}
            GROUP BY day, oi.item_name, oi.price
            ORDER BY day, oi.item_name
        ''', params)
        item_rows = cur.fetchall()
        conn.close()

        # Group items by day
        items_by_day = {}
        for row in item_rows:
            d = row[0]
            items_by_day.setdefault(d, []).append({
                'name':     row[1],
                'category': row[2],
                'quantity': int(row[3]),
                'price':    round(float(row[4]), 2),
                'amount':   round(float(row[5]), 2),
            })

        result = []
        for row in day_rows:
            day_str   = row[0]          # 'YYYY-MM-DD'
            sgst      = round(float(row[1] or 0), 2)
            cgst      = round(float(row[2] or 0), 2)
            day_total = round(float(row[3] or 0), 2)
            items     = items_by_day.get(day_str, [])
            subtotal  = round(sum(it['amount'] for it in items), 2)
            result.append({
                'date':      datetime.date.fromisoformat(day_str),
                'items':     items,
                'subtotal':  subtotal,
                'sgst':      sgst,
                'cgst':      cgst,
                'day_total': day_total,
            })
        return result
    except Exception as e:
        print(f'_get_daily_breakdown error: {e}')
        return []


def _write_monthly_sheet(ws, restaurant: str, day_data: list, ref_dt: datetime.datetime):
    """
    Write the monthly report into worksheet `ws` using the format:

        MY RESTAURANT                                    (cols A–F merged)
        Monthly Sales Report  —  May 2026                (cols A–F merged)
        (blank row)
          Wednesday, 27 May 2026                         (cols A–F merged, indented)
        #  Item Name  Category  Qty  Rate (Rs.)  Amount (Rs.)
        1  …          …         …    …           …
        Subtotal                                 …       (col F)
        SGST  (2.5%)                             …
        CGST  (2.5%)                             …
        Day Total  —  Wednesday, 27 May 2026     …
        (blank row)
          Thursday, 28 May 2026
        … (repeat)

    Columns: A=#, B=Item Name, C=Category, D=Qty, E=Rate (Rs.), F=Amount (Rs.)
    """
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    NCOLS = 6   # A–F

    # ── Colours ────────────────────────────────────────────────────────────────
    HDR_BG    = '1a3c5e'   # deep navy – restaurant name / report title
    WHITE     = 'FFFFFF'
    COL_HDR   = '2d5986'   # column header row per day
    DATE_BG   = 'EAF2FB'   # day date heading background
    ROW_ODD   = 'F5F9FE'   # alternating item row tint
    ROW_EVEN  = 'FFFFFF'
    SUBTOT_BG = 'EDF3FB'   # subtotal / tax rows
    DAYTOT_BG = 'D6E4F0'   # day total row
    FONT_DARK = '1a3c5e'

    # ── Borders ────────────────────────────────────────────────────────────────
    def _thin():
        s = Side(style='thin', color='C5D5E8')
        return Border(left=s, right=s, top=s, bottom=s)

    def _btm_only():
        s = Side(style='thin', color='C5D5E8')
        return Border(bottom=s)

    def _thick():
        s = Side(style='medium', color=HDR_BG)
        return Border(left=s, right=s, top=s, bottom=s)

    # ── Alignments ─────────────────────────────────────────────────────────────
    al_center = Alignment(horizontal='center', vertical='center')
    al_left   = Alignment(horizontal='left',   vertical='center', indent=1)
    al_right  = Alignment(horizontal='right',  vertical='center')

    money_fmt = '#,##0.00'

    # Helper: merge A–F for a row and set common attrs
    def _merged_row(row_num, value, font=None, fill=None, alignment=None, height=None):
        ws.merge_cells(f'A{row_num}:F{row_num}')
        c = ws.cell(row=row_num, column=1)
        c.value     = value
        if font:      c.font      = font
        if fill:      c.fill      = fill
        if alignment: c.alignment = alignment
        if height:    ws.row_dimensions[row_num].height = height
        return c

    # ── Column widths ──────────────────────────────────────────────────────────
    col_widths = [5, 28, 18, 6, 14, 14]   # #, Item, Category, Qty, Rate, Amount
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Row 1: Restaurant name ─────────────────────────────────────────────────
    _merged_row(
        1, restaurant,
        font      = Font(bold=True, size=14, name='Arial', color=WHITE),
        fill      = PatternFill('solid', fgColor=HDR_BG),
        alignment = al_center,
        height    = 24,
    )

    # ── Row 2: Report title ────────────────────────────────────────────────────
    report_title = f'Monthly Sales Report  \u2014  {ref_dt.strftime("%B %Y")}'
    _merged_row(
        2, report_title,
        font      = Font(bold=True, size=11, name='Arial', color=WHITE),
        fill      = PatternFill('solid', fgColor=HDR_BG),
        alignment = al_center,
        height    = 18,
    )

    # ── Row 3: blank spacer ────────────────────────────────────────────────────
    ws.row_dimensions[3].height = 8
    cur_row = 4

    # ── Per-day sections ───────────────────────────────────────────────────────
    for day in day_data:
        d         = day['date']
        items     = day['items']
        subtotal  = day['subtotal']
        sgst      = day['sgst']
        cgst      = day['cgst']
        day_total = day['day_total']

        # Derive GST % label (compute from subtotal if possible)
        if subtotal > 0:
            sgst_pct = round(sgst / subtotal * 100, 2)
            cgst_pct = round(cgst / subtotal * 100, 2)
        else:
            sgst_pct = cgst_pct = 2.5

        day_label  = d.strftime('%A, %d %B %Y')   # "Wednesday, 27 May 2026"
        date_label = f'  {day_label}'              # leading spaces = soft-indent

        # ── Date heading row ──────────────────────────────────────────────
        _merged_row(
            cur_row, date_label,
            font      = Font(bold=True, size=10, name='Arial', color=FONT_DARK),
            fill      = PatternFill('solid', fgColor=DATE_BG),
            alignment = Alignment(horizontal='left', vertical='center', indent=2),
            height    = 18,
        )
        cur_row += 1

        # ── Column header row ─────────────────────────────────────────────
        col_headers = ['#', 'Item Name', 'Category', 'Qty', 'Rate (Rs.)', 'Amount (Rs.)']
        for ci, hdr in enumerate(col_headers, start=1):
            c = ws.cell(row=cur_row, column=ci)
            c.value     = hdr
            c.font      = Font(bold=True, size=9, name='Arial', color=WHITE)
            c.fill      = PatternFill('solid', fgColor=COL_HDR)
            c.alignment = al_center if ci != 2 else Alignment(horizontal='left', vertical='center', indent=1)
            c.border    = _thin()
        ws.row_dimensions[cur_row].height = 16
        cur_row += 1

        # ── Item rows ─────────────────────────────────────────────────────
        for idx, item in enumerate(items, start=1):
            bg   = ROW_ODD if idx % 2 == 1 else ROW_EVEN
            fill = PatternFill('solid', fgColor=bg)
            row  = [idx, item['name'], item['category'], item['quantity'],
                    item['price'], item['amount']]
            for ci, val in enumerate(row, start=1):
                c = ws.cell(row=cur_row, column=ci)
                c.value  = val
                c.fill   = fill
                c.border = _thin()
                c.font   = Font(size=9, name='Arial')
            ws.cell(row=cur_row, column=1).alignment = al_center
            ws.cell(row=cur_row, column=2).alignment = al_left
            ws.cell(row=cur_row, column=3).alignment = al_left
            ws.cell(row=cur_row, column=4).alignment = al_center
            ws.cell(row=cur_row, column=5).alignment = al_right
            ws.cell(row=cur_row, column=5).number_format = money_fmt
            ws.cell(row=cur_row, column=6).alignment = al_right
            ws.cell(row=cur_row, column=6).number_format = money_fmt
            ws.row_dimensions[cur_row].height = 15
            cur_row += 1

        # ── Subtotal row ──────────────────────────────────────────────────
        for ci in range(1, NCOLS + 1):
            c = ws.cell(row=cur_row, column=ci)
            c.fill   = PatternFill('solid', fgColor=SUBTOT_BG)
            c.border = _thin()
            c.font   = Font(size=9, name='Arial')
        ws.merge_cells(f'A{cur_row}:E{cur_row}')
        ws.cell(row=cur_row, column=1).value     = 'Subtotal'
        ws.cell(row=cur_row, column=1).font      = Font(bold=True, size=9, name='Arial', color=FONT_DARK)
        ws.cell(row=cur_row, column=1).alignment = al_left
        ws.cell(row=cur_row, column=6).value          = subtotal
        ws.cell(row=cur_row, column=6).number_format  = money_fmt
        ws.cell(row=cur_row, column=6).alignment      = al_right
        ws.cell(row=cur_row, column=6).font           = Font(bold=True, size=9, name='Arial')
        ws.row_dimensions[cur_row].height = 15
        cur_row += 1

        # ── SGST row ──────────────────────────────────────────────────────
        for ci in range(1, NCOLS + 1):
            c = ws.cell(row=cur_row, column=ci)
            c.fill   = PatternFill('solid', fgColor=SUBTOT_BG)
            c.border = _thin()
            c.font   = Font(size=9, name='Arial')
        ws.merge_cells(f'A{cur_row}:E{cur_row}')
        sgst_label = f'SGST  ({sgst_pct}%)'
        ws.cell(row=cur_row, column=1).value     = sgst_label
        ws.cell(row=cur_row, column=1).alignment = al_left
        ws.cell(row=cur_row, column=6).value         = sgst
        ws.cell(row=cur_row, column=6).number_format = money_fmt
        ws.cell(row=cur_row, column=6).alignment     = al_right
        ws.row_dimensions[cur_row].height = 15
        cur_row += 1

        # ── CGST row ──────────────────────────────────────────────────────
        for ci in range(1, NCOLS + 1):
            c = ws.cell(row=cur_row, column=ci)
            c.fill   = PatternFill('solid', fgColor=SUBTOT_BG)
            c.border = _thin()
            c.font   = Font(size=9, name='Arial')
        ws.merge_cells(f'A{cur_row}:E{cur_row}')
        cgst_label = f'CGST  ({cgst_pct}%)'
        ws.cell(row=cur_row, column=1).value     = cgst_label
        ws.cell(row=cur_row, column=1).alignment = al_left
        ws.cell(row=cur_row, column=6).value         = cgst
        ws.cell(row=cur_row, column=6).number_format = money_fmt
        ws.cell(row=cur_row, column=6).alignment     = al_right
        ws.row_dimensions[cur_row].height = 15
        cur_row += 1

        # ── Day Total row ─────────────────────────────────────────────────
        for ci in range(1, NCOLS + 1):
            c = ws.cell(row=cur_row, column=ci)
            c.fill   = PatternFill('solid', fgColor=DAYTOT_BG)
            c.border = _thick()
            c.font   = Font(bold=True, size=9, name='Arial', color=FONT_DARK)
        ws.merge_cells(f'A{cur_row}:E{cur_row}')
        ws.cell(row=cur_row, column=1).value     = f'Day Total  \u2014  {day_label}'
        ws.cell(row=cur_row, column=1).alignment = al_left
        ws.cell(row=cur_row, column=6).value         = day_total
        ws.cell(row=cur_row, column=6).number_format = money_fmt
        ws.cell(row=cur_row, column=6).alignment     = al_right
        ws.row_dimensions[cur_row].height = 16
        cur_row += 1

        # ── Blank spacer between days ─────────────────────────────────────
        ws.row_dimensions[cur_row].height = 8
        cur_row += 1


# ─── pywebview API ────────────────────────────────────────────────────────────
class Api:
    def close_window(self):
        window.destroy()

    def load_config(self):
        return load_config_from_file()

    def save_config(self, data):
        return save_config_to_file(data)

    def get_session_count(self):
        return load_config_from_file().get('sessionCount', 0)

    # --- License API ---
    def get_system_code(self):
        return get_system_code()

    def activate_license(self, license_key: str):
        sc     = get_system_code()
        result = verify_license_key(license_key, sc)
        if result['valid'] and not result['expired'] and not result.get('clock_tampered'):
            save_license(license_key)
        return result

    def check_license(self):
        sc  = get_system_code()
        key = load_saved_license()
        if not key:
            return {'valid': False, 'expired': False, 'error': 'No license found.', 'system_code': sc, 'clock_tampered': False}
        result = verify_license_key(key, sc)
        result['system_code'] = sc
        return result

    def renew_license(self, new_key: str):
        """Renew/replace an expired license with a new key."""
        return self.activate_license(new_key)

    # --- Excel Sales DB ---
    def get_excel_path(self) -> str:
        """Return the path of the auto-saved sales Excel file."""
        return EXCEL_DB_FILE

    def save_order_to_excel(self, order: dict) -> dict:
        """
        After each completed order, rebuild the entire monthly sales_data.xlsx
        in the requested format: one sheet with all days grouped, each day having
        a date heading, numbered item rows (with category), Subtotal, SGST, CGST,
        and Day Total.  The whole file is rebuilt from SQLite so it's always fresh.
        Returns {'ok': True} or {'ok': False, 'error': '...'}.
        """
        if not _OPENPYXL_OK:
            return {'ok': False, 'error': 'openpyxl not installed. Run: pip install openpyxl'}

        try:
            r = load_settings_from_db()
            cfg        = r['settings'] if r['ok'] and r['settings'] else load_config_from_file()
            restaurant = cfg.get('restaurantName', 'MY RESTAURANT')

            # ── Fetch all orders grouped by day from SQLite ─────────────────
            order_dt = datetime.datetime.fromisoformat(
                order.get('date', datetime.datetime.now().isoformat())
            )
            month_start = order_dt.strftime('%Y-%m-01') + 'T00:00:00'
            month_end   = order_dt.strftime('%Y-%m-') + '31T23:59:59'

            day_data = _get_daily_breakdown(month_start, month_end)

            # ── Build workbook ──────────────────────────────────────────────
            wb = Workbook()
            ws = wb.active
            ws.title = order_dt.strftime('%B %Y')   # e.g. "May 2026"

            _write_monthly_sheet(ws, restaurant, day_data, order_dt)

            wb.save(EXCEL_DB_FILE)
            return {'ok': True, 'file': EXCEL_DB_FILE}

        except Exception as e:
            return {'ok': False, 'error': str(e)}

    def load_orders_from_excel(self) -> dict:
        """
        Load all orders from the persistent sales_data.xlsx file.
        Returns {'ok': True, 'orders': [...]} or {'ok': False, 'error': '...', 'orders': []}.
        """
        if not _OPENPYXL_OK:
            return {'ok': False, 'error': 'openpyxl not installed', 'orders': []}

        orders = []
        try:
            if not os.path.exists(EXCEL_DB_FILE):
                return {'ok': True, 'orders': []}

            wb = load_workbook(EXCEL_DB_FILE)
            
            # Load orders from all sheets (one sheet per day)
            for sheet_name in wb.sheetnames:
                try:
                    ws = wb[sheet_name]
                    # Skip non-data rows (banner, header, etc.)
                    # Data starts at row 5
                    for row_idx, row in enumerate(ws.iter_rows(min_row=5, values_only=True), start=5):
                        if not row[0] or row[0] == 'Date':  # Skip empty/header rows
                            continue
                        
                        try:
                            # Parse date and time
                            date_str = str(row[0]) if row[0] else ''
                            time_str = str(row[1]) if row[1] else '00:00 AM'
                            datetime_str = f"{date_str} {time_str}"
                            
                            # Parse datetime - try multiple formats
                            order_dt = None
                            for fmt in ['%d-%b-%Y %I:%M %p', '%Y-%m-%d %H:%M:%S', '%d-%b-%Y %H:%M:%S']:
                                try:
                                    order_dt = datetime.datetime.strptime(datetime_str, fmt)
                                    break
                                except:
                                    continue
                            
                            if not order_dt:
                                order_dt = datetime.datetime.now()
                            
                            # Parse amounts
                            subtotal = float(row[4] or 0)
                            discount = float(row[5] or 0)
                            sgst = float(row[6] or 0)
                            cgst = float(row[7] or 0)
                            total = float(row[8] or 0)
                            
                            order = {
                                'id': str(row[2]) if row[2] else '',
                                'date': order_dt.isoformat(),
                                'items': [],  # Excel doesn't store full item details, only summary
                                'subtotal': subtotal,
                                'discount': discount,
                                'sgst': sgst,
                                'cgst': cgst,
                                'gst': sgst,  # For compatibility
                                'total': total
                            }
                            orders.append(order)
                        except Exception as row_err:
                            print(f"Error parsing row {row_idx} in sheet '{sheet_name}': {row_err}")
                            continue
                except Exception as sheet_err:
                    print(f"Error reading sheet '{sheet_name}': {sheet_err}")
                    continue
            
            return {'ok': True, 'orders': orders}

        except Exception as e:
            return {'ok': False, 'error': str(e), 'orders': []}

    def save_order(self, order: dict) -> dict:
        """Save order to SQLite database."""
        return save_order_to_db(order)

    def load_orders(self) -> dict:
        """Load all orders from SQLite database."""
        return load_orders_from_db()

    # --- Settings SQLite API ---
    def save_settings(self, settings: dict) -> dict:
        """Persist settings to SQLite (and config.json for legacy compat)."""
        save_config_to_file(settings)          # keep config.json in sync
        return save_settings_to_db(settings)

    def load_settings(self) -> dict:
        """Load settings from SQLite; fall back to config.json."""
        result = load_settings_from_db()
        if result['ok'] and result['settings']:
            return result
        # Fall back to config.json (first-run or migration)
        r = load_settings_from_db(); cfg = r['settings'] if r['ok'] and r['settings'] else load_config_from_file()
        cfg.pop('sessionCount', None)
        return {'ok': True, 'settings': cfg}

    # --- Cart SQLite API ---
    def save_cart(self, cart: list) -> dict:
        """Persist cart to SQLite."""
        return save_cart_to_db(cart)

    def load_cart(self) -> dict:
        """Load cart from SQLite."""
        return load_cart_from_db()

    # --- Menu & Categories SQLite API ---
    def save_menu(self, menu: list) -> dict:
        """Persist full menu to SQLite."""
        return save_menu_to_db(menu)

    def load_menu(self) -> dict:
        """Load full menu from SQLite."""
        return load_menu_from_db()

    def save_categories(self, categories: list) -> dict:
        """Persist full categories list to SQLite."""
        return save_categories_to_db(categories)

    def load_categories(self) -> dict:
        """Load full categories list from SQLite."""
        return load_categories_from_db()

    def get_sales_summary_api(self, from_date=None, to_date=None) -> dict:
        """Get sales summary with items and revenue."""
        return get_sales_summary(from_date, to_date)

    def print_receipt(self, receipt: dict) -> dict:
        """
        Print a thermal receipt via win32print / win32ui (GDI).
        `receipt` is a dict passed from JS with these fields:
            restaurantName, address, phone, gstin,
            invoiceNo, dateStr, currency,
            items:   [{name, qty, price, total}]
            baseAmount, taxMode, gstRate, sgstRate, cgstRate,
            gstAmount, sgstAmount, cgstAmount,
            discountRate, discountAmount, grandTotal
        Returns {'ok': True} or {'ok': False, 'error': '...'}.
        """
        try:
            import win32print
            import win32ui
            from win32con import (LOGPIXELSX, LOGPIXELSY,
                                  DM_PAPERWIDTH, DM_PAPERLENGTH,
                                  DMORIENT_PORTRAIT, DM_ORIENTATION)
        except ImportError:
            return {'ok': False, 'error': 'win32print not available. Run: pip install pywin32'}

        try:
            # ── Printer & DC ────────────────────────────────────────────────
            printer_name = win32print.GetDefaultPrinter()
            hPrinter     = win32print.OpenPrinter(printer_name)
            pDev         = win32print.GetPrinter(hPrinter, 2)['pDevMode']

            # Thermal paper: 80 mm wide = 3071 units (in 0.1 mm units = 800)
            PAPER_W_MM = 80
            try:
                pDev.PaperWidth  = PAPER_W_MM * 10   # 0.1mm units
                pDev.PaperLength = 0                  # continuous
                pDev.Orientation = DMORIENT_PORTRAIT
                pDev.Fields     |= DM_PAPERWIDTH | DM_PAPERLENGTH | DM_ORIENTATION
            except Exception:
                pass
            win32print.ClosePrinter(hPrinter)

            dc = win32ui.CreateDC()
            dc.CreatePrinterDC(printer_name)
            dc.StartDoc('HBILLSOFT Receipt')
            dc.StartPage()

            # ── DPI & measurements ──────────────────────────────────────────
            dpi_x = dc.GetDeviceCaps(LOGPIXELSX)
            dpi_y = dc.GetDeviceCaps(LOGPIXELSY)

            # Paper width in pixels (use 80 mm / 25.4 * dpi, capped to page width)
            page_w_px = dc.GetDeviceCaps(110)   # HORZRES
            receipt_w = min(int(PAPER_W_MM / 25.4 * dpi_x), page_w_px)

            MARGIN_L = int(0.05 * dpi_x)   # 0.05 inch left margin
            MARGIN_T = int(0.1  * dpi_y)   # 0.1 inch top margin
            LINE_H_N = int(0.18 * dpi_y)   # normal line height
            LINE_H_B = int(0.20 * dpi_y)   # bold line height
            PRINT_W  = receipt_w - MARGIN_L * 2

            # ── Font factory ────────────────────────────────────────────────
            def make_font(size_pt, bold=False):
                height = -int(size_pt * dpi_y / 72)
                weight = 700 if bold else 400
                f = win32ui.CreateFont({
                    'name':   'Courier New',
                    'height': height,
                    'weight': weight,
                    'charset': 0,
                })
                return f

            # ── Text helpers ────────────────────────────────────────────────
            curr = receipt.get('currency', 'Rs.')
            LINE_LEN = 42   # chars at 9pt across 80mm

            def pad_center(s, n=LINE_LEN):
                s = s[:n]; pad = max(0, (n - len(s)) // 2)
                return ' ' * pad + s

            def pad_lr(left, right, n=LINE_LEN):
                left  = left[:n - len(right) - 1]
                space = n - len(left) - len(right)
                return left + ' ' * max(1, space) + right

            def draw_line(text, y, bold=False, size=9, center=False):
                font = make_font(size, bold)
                dc.SelectObject(font)
                if center:
                    text = pad_center(text)
                dc.TextOut(MARGIN_L, y, text)
                return y + (LINE_H_B if bold else LINE_H_N)

            def draw_separator(y, char='='):
                return draw_line(char * LINE_LEN, y)

            # ── Build receipt ───────────────────────────────────────────────
            y = MARGIN_T

            # Header
            y = draw_line(receipt.get('restaurantName', 'MY RESTAURANT'), y, bold=True, size=12, center=True)
            if receipt.get('address'):
                y = draw_line(receipt['address'], y, center=True)
            if receipt.get('phone'):
                y = draw_line('Ph: ' + receipt['phone'], y, center=True)
            if receipt.get('gstin'):
                y = draw_line('GSTIN: ' + receipt['gstin'], y, center=True)

            y = draw_separator(y)

            y = draw_line(pad_center('Date: ' + receipt.get('dateStr', '')), y)
            y = draw_line(pad_center('Invoice: ' + receipt.get('invoiceNo', '')), y, bold=True)

            y = draw_separator(y, '-')

            # Column header
            y = draw_line(pad_lr('Item', f'Qty  {"Rate":>7}  {"Amt":>8}'), y, bold=True)
            y = draw_separator(y, '-')

            # Items
            for item in receipt.get('items', []):
                name  = item.get('name', '')
                qty   = item.get('qty', 1)
                price = item.get('price', 0)
                total = item.get('total', qty * price)
                right = f"{qty}x{curr}{price:.0f}  {curr}{total:.2f}"
                max_name = LINE_LEN - len(right) - 1
                if len(name) > max_name:
                    name = name[:max_name]
                y = draw_line(pad_lr(name, right), y)

            y = draw_separator(y, '-')

            # Totals block
            base   = receipt.get('baseAmount', 0)
            tax_m  = receipt.get('taxMode', 'split')
            y = draw_line(pad_lr('Subtotal', f'{curr}{base:.2f}'), y)

            if tax_m == 'gst' and receipt.get('gstRate', 0) > 0:
                y = draw_line(pad_lr(f"GST incl. ({receipt['gstRate']}%)", f"{curr}{receipt.get('gstAmount',0):.2f}"), y)
            elif tax_m == 'split':
                sr = receipt.get('sgstRate', 0); cr = receipt.get('cgstRate', 0)
                if sr > 0:
                    y = draw_line(pad_lr(f'SGST incl. ({sr}%)', f"{curr}{receipt.get('sgstAmount',0):.2f}"), y)
                if cr > 0:
                    y = draw_line(pad_lr(f'CGST incl. ({cr}%)', f"{curr}{receipt.get('cgstAmount',0):.2f}"), y)

            dr = receipt.get('discountRate', 0)
            if dr > 0:
                y = draw_line(pad_lr(f'Discount ({dr}%)', f"-{curr}{receipt.get('discountAmount',0):.2f}"), y)

            y = draw_separator(y)
            y = draw_line(pad_lr('TOTAL:', f"{curr}{receipt.get('grandTotal',0):.2f}"), y, bold=True, size=11)
            y = draw_separator(y)

            # Footer
            y = draw_line('', y)
            y = draw_line('Thank you! Please visit again.', y, bold=True, center=True)
            y = draw_line('', y)

            dc.EndPage()
            dc.EndDoc()
            dc.DeleteDC()

            return {'ok': True}

        except Exception as e:
            return {'ok': False, 'error': str(e)}

    def export_sales_excel(self, from_date=None, to_date=None) -> dict:
        """
        Build a clean sales report Excel file and save it to the desktop.
        Returns {'ok': True, 'file': path} or {'ok': False, 'error': '...'}.
        """
        if not _OPENPYXL_OK:
            return {'ok': False, 'error': 'openpyxl not installed. Run: pip install openpyxl'}

        try:
            r = load_settings_from_db()
            cfg        = r['settings'] if r['ok'] and r['settings'] else load_config_from_file()
            restaurant = cfg.get('restaurantName', 'MY RESTAURANT')

            day_data = _get_daily_breakdown(from_date, to_date)

            # Determine report period label for the title row
            if from_date or to_date:
                fd_str = (from_date or '').split('T')[0]
                td_str = (to_date   or '').split('T')[0]
                # Parse to a friendly month label if they span one month
                try:
                    ref_dt = datetime.datetime.strptime(fd_str or td_str, '%Y-%m-%d')
                except Exception:
                    ref_dt = datetime.datetime.now()
            else:
                ref_dt = datetime.datetime.now()

            wb = Workbook()
            ws = wb.active
            ws.title = ref_dt.strftime('%B %Y')

            _write_monthly_sheet(ws, restaurant, day_data, ref_dt)

            # ── Save to Desktop ─────────────────────────────────────────────
            today_str = datetime.date.today().strftime('%Y-%m-%d')
            desktop = os.path.join(os.path.expanduser('~'), 'Desktop')
            if not os.path.isdir(desktop):
                desktop = os.path.expanduser('~')
            out_path = os.path.join(desktop, f'HBILLSOFT_Sales_{today_str}.xlsx')

            wb.save(out_path)
            return {'ok': True, 'file': out_path}

        except Exception as e:
            return {'ok': False, 'error': str(e)}

# ─── Startup ──────────────────────────────────────────────────────────────────
api = Api()

# Initialize database
init_database()

config = load_config_from_file()
config['sessionCount'] = config.get('sessionCount', 0) + 1
save_config_to_file(config)

html_file = resource_path('RestoPOS.html')

window = webview.create_window(
    title='HBILLSOFT',
    url='file:///' + html_file.replace('\\', '/'),
    width=1280,
    height=800,
    min_size=(1024, 600),
    resizable=True,
    fullscreen=True,
    js_api=api
)

webview.start()
